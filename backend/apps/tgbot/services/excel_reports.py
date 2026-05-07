"""
Генерация Excel-отчётов для ежедневной рассылки в Telegram.

Два отчёта, оба отправляются в 22:00:
    1. stock_balance_xlsx — остатки по всем складам и SKU
    2. debtors_xlsx       — список дебиторов (контрагентов с долгом)

Файлы шлёт `apps.tgbot.bot.send_document` через sendDocument API.
Имя файла формата: <YYYY-MM-DD>_<тип>.xlsx
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ─── Стили ──────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill("solid", fgColor="E8751A")  # бренд-orange
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TOTAL_FONT = Font(bold=True, size=11)
_TOTAL_FILL = PatternFill("solid", fgColor="FEF3C7")
_NUMBER_FORMAT = "#,##0.00"
_INT_FORMAT = "#,##0"


def _autofit(ws, max_widths: dict[str, int] | None = None) -> None:
    """Грубый autofit: ширина колонки = max(len(value)) + 2, capped 50."""
    for col_idx, col in enumerate(ws.columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            length = len(str(v))
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(
            max_len + 2,
            (max_widths or {}).get(letter, 50),
        )


def _write_header(ws, headers: list[str]) -> None:
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


# ─── Stock balances ─────────────────────────────────────────────────────


def generate_stock_balance_xlsx(organization, *, today: date | None = None) -> bytes:
    """Excel-отчёт остатков сырья и готовой продукции по всем складам.

    Источник: StockMovement aggregated по nomenclature × warehouse_to/from.
    Колонки:
        Склад | Модуль | SKU | Наименование | Ед | Σ Приход | Σ Расход | Остаток
    Сортировка: склад → SKU.
    """
    from apps.warehouses.models import StockMovement, Warehouse

    today = today or date.today()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Остатки {today.isoformat()}"

    headers = [
        "Склад", "Модуль", "SKU", "Наименование", "Ед",
        "Σ Приход", "Σ Расход", "Остаток",
    ]
    _write_header(ws, headers)

    # Считаем по каждой паре (warehouse, nomenclature) сумму incoming/outgoing.
    # Остаток = Σ incoming − Σ outgoing − Σ write_off − Σ shrinkage.
    # Transfer не вычитаем (он одновременно in и out на разные склады).
    movements = (
        StockMovement.objects
        .filter(organization=organization)
        .select_related("nomenclature", "nomenclature__unit",
                        "warehouse_from", "warehouse_to", "module")
    )

    # Aggregate в питоне — модель просто, проще обработать руками.
    # key = (warehouse_id, nomenclature_id) → {wh, nom, in, out, mod}
    bag: dict[tuple, dict] = {}

    def _entry(wh: Warehouse | None, nom):
        if wh is None:
            return None
        key = (wh.id, nom.id)
        e = bag.get(key)
        if e is None:
            e = {
                "warehouse": wh,
                "nomenclature": nom,
                "module": wh.module if wh.module_id else None,
                "incoming": Decimal("0"),
                "outgoing": Decimal("0"),
            }
            bag[key] = e
        return e

    for m in movements.iterator():
        nom = m.nomenclature
        qty = Decimal(m.quantity or 0)
        if m.kind == StockMovement.Kind.INCOMING:
            e = _entry(m.warehouse_to, nom)
            if e:
                e["incoming"] += qty
        elif m.kind in (
            StockMovement.Kind.OUTGOING,
            StockMovement.Kind.WRITE_OFF,
            StockMovement.Kind.SHRINKAGE,
        ):
            e = _entry(m.warehouse_from, nom)
            if e:
                e["outgoing"] += qty
        elif m.kind == StockMovement.Kind.TRANSFER:
            ein = _entry(m.warehouse_to, nom)
            eout = _entry(m.warehouse_from, nom)
            if ein:
                ein["incoming"] += qty
            if eout:
                eout["outgoing"] += qty

    rows = sorted(
        bag.values(),
        key=lambda e: (
            e["warehouse"].code if e["warehouse"] else "",
            e["nomenclature"].sku or "",
        ),
    )

    row_idx = 2
    total_balance = Decimal("0")
    for e in rows:
        balance = e["incoming"] - e["outgoing"]
        if balance == 0 and e["incoming"] == 0 and e["outgoing"] == 0:
            continue  # совсем пустые не показываем
        ws.cell(row=row_idx, column=1, value=e["warehouse"].code)
        ws.cell(row=row_idx, column=2,
                value=e["module"].code if e["module"] else "—")
        ws.cell(row=row_idx, column=3, value=e["nomenclature"].sku)
        ws.cell(row=row_idx, column=4, value=e["nomenclature"].name)
        ws.cell(row=row_idx, column=5,
                value=getattr(e["nomenclature"].unit, "code", "—"))
        ws.cell(row=row_idx, column=6, value=float(e["incoming"]))
        ws.cell(row=row_idx, column=7, value=float(e["outgoing"]))
        bal_cell = ws.cell(row=row_idx, column=8, value=float(balance))
        bal_cell.font = Font(bold=True)
        if balance < 0:
            bal_cell.font = Font(bold=True, color="C0392B")
        for col in (6, 7, 8):
            ws.cell(row=row_idx, column=col).number_format = _NUMBER_FORMAT
        total_balance += balance
        row_idx += 1

    # Итог
    if row_idx > 2:
        ws.cell(row=row_idx, column=1, value="Итого").font = _TOTAL_FONT
        cell = ws.cell(row=row_idx, column=8, value=float(total_balance))
        cell.font = _TOTAL_FONT
        cell.number_format = _NUMBER_FORMAT
        for col_idx in range(1, 9):
            ws.cell(row=row_idx, column=col_idx).fill = _TOTAL_FILL

    _autofit(ws, max_widths={"D": 50})

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ─── Debtors list ───────────────────────────────────────────────────────


def generate_debtors_xlsx(organization, *, today: date | None = None) -> bytes:
    """Excel-отчёт списка дебиторов с aging-разбивкой.

    Колонки:
        Код | Контрагент | Текущ. | 0-30 | 31-60 | 61-90 | 90+ | Всего | Просрочка дн.
    Сортировка: по убыванию total.
    """
    from apps.sales.services.aging import compute_aging_report

    today = today or date.today()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Должники {today.isoformat()}"

    headers = [
        "Код", "Контрагент",
        "Текущие, UZS", "0-30 дн", "31-60 дн", "61-90 дн", "90+ дн",
        "Всего долг, UZS", "Старая просрочка, дн",
    ]
    _write_header(ws, headers)

    report = compute_aging_report(organization, today=today)
    row_idx = 2
    total = Decimal("0")
    for r in report.rows:
        ws.cell(row=row_idx, column=1, value=r.code)
        ws.cell(row=row_idx, column=2, value=r.name)
        ws.cell(row=row_idx, column=3, value=float(r.current))
        ws.cell(row=row_idx, column=4, value=float(r.b_0_30))
        ws.cell(row=row_idx, column=5, value=float(r.b_31_60))
        ws.cell(row=row_idx, column=6, value=float(r.b_61_90))
        ws.cell(row=row_idx, column=7, value=float(r.b_90_plus))
        total_cell = ws.cell(row=row_idx, column=8, value=float(r.total))
        total_cell.font = Font(bold=True)
        ws.cell(row=row_idx, column=9, value=r.oldest_overdue_days)
        for col in (3, 4, 5, 6, 7, 8):
            ws.cell(row=row_idx, column=col).number_format = _NUMBER_FORMAT
        # Подсветка строки если просрочка > 30 дней
        if r.oldest_overdue_days > 30:
            highlight = PatternFill("solid", fgColor="FEE2E2")
            for col_idx in range(1, 10):
                ws.cell(row=row_idx, column=col_idx).fill = highlight
        total += Decimal(r.total)
        row_idx += 1

    # Итог
    if row_idx > 2:
        ws.cell(row=row_idx, column=1, value="Итого").font = _TOTAL_FONT
        cell = ws.cell(row=row_idx, column=8, value=float(total))
        cell.font = _TOTAL_FONT
        cell.number_format = _NUMBER_FORMAT
        for col_idx in range(1, 10):
            ws.cell(row=row_idx, column=col_idx).fill = _TOTAL_FILL

    _autofit(ws, max_widths={"B": 50})

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


# ─── Имя файла по правилам пользователя ─────────────────────────────────

def stock_filename(today: date | None = None) -> str:
    today = today or date.today()
    return f"{today.isoformat()}_otchet_o_sklade.xlsx"


def debtors_filename(today: date | None = None) -> str:
    today = today or date.today()
    return f"{today.isoformat()}_spisok_doljnikov.xlsx"
